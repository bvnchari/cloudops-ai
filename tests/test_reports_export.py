import os

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from core.export import export_excel, export_pdf
from core.kpi import KPIEngine
from core.report_periods import ReportPeriod, build_bundle, build_period
from core.scheduler import CRON_MAP, ReportSubscription, SMTPConfig
from pipeline import run_pipeline


@pytest.fixture(scope="module")
def pipeline_result():
    return run_pipeline(verbose=False)


@pytest.fixture(scope="module")
def sample_bundle(pipeline_result):
    from core.insights import service_scorecard
    r = pipeline_result
    period = build_period("weekly")
    return build_bundle(period=period, kpi=r["kpi"], stats=r["stats"],
                        scorecard=service_scorecard(r["incidents"]))


# ---------------------------------------------------------------- periods --

def test_build_period_known_types():
    for ptype, days in [("daily", 1), ("weekly", 7), ("monthly", 30), ("quarterly", 90)]:
        p = build_period(ptype)
        assert p.days == days
        assert p.period_type == ptype
        assert p.end > p.start


def test_build_period_custom_requires_days():
    with pytest.raises(ValueError):
        build_period("custom")
    p = build_period("custom", custom_days=45)
    assert p.days == 45


def test_build_period_unknown_type_raises():
    with pytest.raises(ValueError):
        build_period("fortnightly")


def test_build_bundle_defaults_empty_lists():
    period = build_period("daily")
    kpi = KPIEngine().compute(10, [])
    bundle = build_bundle(period=period, kpi=kpi, stats={"raw_alerts": 10,
                          "incidents": 0, "deduped_events": 0, "noise_reduction_pct": 0})
    assert bundle.slo_statuses == []
    assert bundle.scorecard == []


# ---------------------------------------------------------------- export --

def test_export_pdf_creates_valid_file(tmp_path, pipeline_result):
    r = pipeline_result
    period = build_period("monthly")
    bundle = build_bundle(period=period, kpi=r["kpi"], stats=r["stats"])
    out = tmp_path / "report.pdf"
    path = export_pdf(bundle, str(out))
    assert os.path.exists(path)
    reader = PdfReader(path)
    assert len(reader.pages) >= 1
    text = reader.pages[0].extract_text()
    assert "Reliability Report" in text


def test_export_excel_creates_valid_workbook(tmp_path, pipeline_result):
    r = pipeline_result
    period = build_period("quarterly")
    bundle = build_bundle(period=period, kpi=r["kpi"], stats=r["stats"])
    out = tmp_path / "report.xlsx"
    path = export_excel(bundle, str(out))
    assert os.path.exists(path)
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "KPIs" in wb.sheetnames
    kpi_ws = wb["KPIs"]
    assert kpi_ws["A1"].value == "Metric"


def test_export_handles_empty_slo_sla_gracefully(tmp_path):
    period = build_period("daily")
    kpi = KPIEngine().compute(5, [])
    bundle = build_bundle(period=period, kpi=kpi,
                          stats={"raw_alerts": 5, "incidents": 0,
                                "deduped_events": 0, "noise_reduction_pct": 0})
    pdf_path = export_pdf(bundle, str(tmp_path / "empty.pdf"))
    xlsx_path = export_excel(bundle, str(tmp_path / "empty.xlsx"))
    assert os.path.exists(pdf_path)
    assert os.path.exists(xlsx_path)


# -------------------------------------------------------------- scheduler --

def test_cron_map_covers_all_standard_periods():
    for ptype in ("daily", "weekly", "monthly", "quarterly"):
        assert ptype in CRON_MAP


def test_smtp_config_requires_env(monkeypatch):
    for var in ("SMTP_HOST", "SMTP_USER", "SMTP_PASSWORD"):
        monkeypatch.delenv(var, raising=False)
    with pytest.raises(RuntimeError):
        SMTPConfig.from_env()


def test_smtp_config_from_env(monkeypatch):
    monkeypatch.setenv("SMTP_HOST", "smtp.example.com")
    monkeypatch.setenv("SMTP_USER", "reports@example.com")
    monkeypatch.setenv("SMTP_PASSWORD", "secret")
    cfg = SMTPConfig.from_env()
    assert cfg.host == "smtp.example.com"
    assert cfg.port == 587
    assert cfg.sender == "reports@example.com"


def test_report_subscription_defaults():
    sub = ReportSubscription(name="cfo-weekly", recipients=["cfo@example.com"])
    assert sub.period_type == "weekly"
    assert sub.formats == ("pdf",)
    assert sub.active is True
